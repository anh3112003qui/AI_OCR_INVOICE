from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from invoice_ocr.models import InvoiceData
from invoice_ocr.validation import ValidationResult

_SHEET_NAME = "29a_AI_OCR_STAGING"
_HEADER_ROW = 36
_DATA_START_ROW = 37
# Column A has auto-STT formula — skip it; data starts at column B (index 2)
_DATA_COL_START = 2   # B = Ngày HĐ


def _find_next_empty_row(sheet) -> int:
    """Find first empty row by checking column B (Ngày HĐ) downward from DATA_START_ROW."""
    row = _DATA_START_ROW
    while sheet.cell(row=row, column=_DATA_COL_START).value not in (None, ""):
        row += 1
    return row


class ExcelExporter:
    """
    Appends validated InvoiceData rows into sheet 29a_AI_OCR_STAGING of an existing workbook.

    Writes columns B–Q (16 AI-supplied data columns).
    Column A (STT) uses an existing sheet formula and is not touched.
    Columns R onward (lookup results, VAT totals, Validation) are left untouched.
    """

    def __init__(self, workbook_path: str | Path) -> None:
        self._path = Path(workbook_path)
        if not self._path.exists():
            raise FileNotFoundError(f"Workbook not found: {self._path}")

    def append(self, result: ValidationResult) -> int:
        """
        Append all line items from a ValidationResult into the staging sheet.

        Args:
            result: output from InvoiceValidator containing normalized InvoiceData

        Returns:
            number of rows written
        """
        wb = load_workbook(self._path)

        if _SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Sheet '{_SHEET_NAME}' not found in workbook.")

        sheet = wb[_SHEET_NAME]
        next_row = _find_next_empty_row(sheet)
        invoice = result.invoice
        rows_written = 0

        for item in invoice.line_items:
            # 16 data columns mapped to B(2)..Q(17)
            row_data = [
                self._as_datetime(invoice.ngay_hd),  # B — Ngày HĐ
                invoice.so_hd,                        # C — Số HĐ
                invoice.ma_ncc,                       # D — Mã NCC
                invoice.mst_ncc,                      # E — MST NCC
                invoice.ten_ncc_raw,                  # F — Tên NCC (raw)
                item.ma_vt,                           # G — Mã VT
                item.ten_vt_raw,                      # H — Tên VT (raw)
                item.dvt,                             # I — ĐVT
                item.so_luong,                        # J — Số lượng
                self._as_float(item.don_gia),         # K — Đơn giá
                item.ma_thue,                         # L — Mã Thuế
                item.thue_suat_pct,                   # M — Thuế suất %
                item.tk_kho,                          # N — TK kho
                item.ma_bp,                           # O — Mã BP
                item.ma_da,                           # P — Mã DA
                item.ghi_chu,                         # Q — Ghi chú
            ]

            for col_offset, value in enumerate(row_data):
                cell = sheet.cell(
                    row=next_row,
                    column=_DATA_COL_START + col_offset,
                    value=value,
                )
                cell.alignment = Alignment(wrap_text=False)

            next_row += 1
            rows_written += 1

        wb.save(self._path)
        return rows_written

    def _as_datetime(self, d: date | None) -> datetime | None:
        """Convert date to datetime so openpyxl renders it as a date cell."""
        if d is None:
            return None
        return datetime(d.year, d.month, d.day)

    def _as_float(self, value: Decimal | None) -> float | None:
        if value is None:
            return None
        return float(value)