from __future__ import annotations

import shutil
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from invoice_ocr.exporters.excel_exporter import ExcelExporter
from invoice_ocr.models import InvoiceData, InvoiceLineItem
from invoice_ocr.validation import InvoiceValidator

_WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "data" / "HE_THONG_KE_TOAN_TT99.xlsx"
_OUTPUT_PATH = Path(__file__).resolve().parent.parent / "data" / "output" / "HE_THONG_KE_TOAN_TT99_ocr_out.xlsx"
_SHEET_NAME = "29a_AI_OCR_STAGING"
_DATA_START_ROW = 37


def _make_validated_invoice() -> object:
    """Build a ValidationResult from the sample invoice used in extractor tests."""
    invoice = InvoiceData(
        ngay_hd=date(2026, 4, 12),
        so_hd="0000087",
        mst_ncc="0310987654",
        ten_ncc_raw="CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ AN PHÁT",
        ma_ncc=None,
        line_items=[
            InvoiceLineItem(
                ten_vt_raw="Hạt nhựa PP copolymer (PP-C30S)",
                dvt="KG",
                so_luong=3500,
                don_gia=Decimal("38000"),
                thue_suat_pct=8,
            ),
            InvoiceLineItem(
                ten_vt_raw="Hạt nhựa HDPE (HD5502XA)",
                dvt="KG",
                so_luong=1200,
                don_gia=Decimal("42500"),
                thue_suat_pct=8,
            ),
            InvoiceLineItem(
                ten_vt_raw="Chất phụ gia chống UV (UV-328)",
                dvt="KG",
                so_luong=80,
                don_gia=Decimal("185000"),
                thue_suat_pct=8,
            ),
        ],
    )
    return InvoiceValidator().validate(invoice)


def _count_data_rows(path: Path) -> int:
    wb = load_workbook(path)
    sheet = wb[_SHEET_NAME]
    count = 0
    row = _DATA_START_ROW
    while sheet.cell(row=row, column=2).value not in (None, ""):
        count += 1
        row += 1
    return count


def test_excel_exporter_appends_rows(tmp_path: Path) -> None:
    """Smoke test: ExcelExporter appends correct number of rows to the staging sheet."""
    workbook_copy = tmp_path / "test_wb.xlsx"
    shutil.copy(_WORKBOOK_PATH, workbook_copy)

    rows_before = _count_data_rows(workbook_copy)
    result = _make_validated_invoice()

    exporter = ExcelExporter(workbook_copy)
    written = exporter.append(result)

    assert written == 3, f"Expected 3 rows written, got {written}"

    rows_after = _count_data_rows(workbook_copy)
    assert rows_after == rows_before + 3, (
        f"Expected {rows_before + 3} rows after append, got {rows_after}"
    )


def test_excel_exporter_data_integrity(tmp_path: Path) -> None:
    """Verify cell values of the first appended row are correct."""
    workbook_copy = tmp_path / "test_wb.xlsx"
    shutil.copy(_WORKBOOK_PATH, workbook_copy)

    rows_before = _count_data_rows(workbook_copy)
    result = _make_validated_invoice()

    ExcelExporter(workbook_copy).append(result)

    wb = load_workbook(workbook_copy)
    sheet = wb[_SHEET_NAME]
    first_new_row = _DATA_START_ROW + rows_before

    ngay_hd   = sheet.cell(row=first_new_row, column=2).value
    so_hd     = sheet.cell(row=first_new_row, column=3).value
    mst_ncc   = sheet.cell(row=first_new_row, column=5).value
    ten_ncc   = sheet.cell(row=first_new_row, column=6).value
    ten_vt    = sheet.cell(row=first_new_row, column=8).value
    dvt       = sheet.cell(row=first_new_row, column=9).value
    so_luong  = sheet.cell(row=first_new_row, column=10).value
    don_gia   = sheet.cell(row=first_new_row, column=11).value
    thue_suat = sheet.cell(row=first_new_row, column=13).value

    print(f"\n--- First appended row (row {first_new_row}) ---")
    print(f"Ngày HĐ : {ngay_hd}")
    print(f"Số HĐ   : {so_hd}")
    print(f"MST NCC : {mst_ncc}")
    print(f"Tên NCC : {ten_ncc}")
    print(f"Tên VT  : {ten_vt}")
    print(f"ĐVT     : {dvt}")
    print(f"SL      : {so_luong}")
    print(f"Đơn giá : {don_gia}")
    print(f"Thuế %  : {thue_suat}")

    assert isinstance(ngay_hd, datetime), f"ngay_hd should be datetime, got {type(ngay_hd)}"
    assert ngay_hd.date() == date(2026, 4, 12)
    assert so_hd == "0000087"
    assert mst_ncc == "0310987654"
    assert ten_ncc == "CÔNG TY TNHH THƯƠNG MẠI VÀ DỊCH VỤ AN PHÁT"
    assert ten_vt == "Hạt nhựa PP copolymer (PP-C30S)"
    assert dvt == "KG"
    assert so_luong == 3500
    assert don_gia == 38000.0
    assert thue_suat == 8


def test_excel_exporter_append_twice(tmp_path: Path) -> None:
    """Verify second append starts after first batch, not overwriting it."""
    workbook_copy = tmp_path / "test_wb.xlsx"
    shutil.copy(_WORKBOOK_PATH, workbook_copy)

    rows_before = _count_data_rows(workbook_copy)
    result = _make_validated_invoice()
    exporter = ExcelExporter(workbook_copy)

    exporter.append(result)
    exporter.append(result)

    rows_after = _count_data_rows(workbook_copy)
    assert rows_after == rows_before + 6


def test_excel_exporter_save_output_for_review() -> None:
    """
    Write output to data/output/ for manual review in Excel.
    Not a correctness test — always passes.
    """
    _OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(_WORKBOOK_PATH, _OUTPUT_PATH)

    result = _make_validated_invoice()
    exporter = ExcelExporter(_OUTPUT_PATH)
    written = exporter.append(result)

    print(f"\n--- Output saved for review ---")
    print(f"File : {_OUTPUT_PATH}")
    print(f"Rows written: {written}")
    print(f"Open sheet  : {_SHEET_NAME}")